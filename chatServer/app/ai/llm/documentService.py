"""
Document Service
Handles decoding base64 vault documents (PDF, CSV, XLSX),
extracting text, and chunking into RAG-ready pieces.

Rename from pdfService.py → documentService.py.
Update import in ragPipeline.py accordingly.
"""
import base64
import csv
import io
import logging
from dataclasses import dataclass
from typing import List, Optional

import fitz  # pymupdf — PDF text rendering for OCR
import pdfplumber # For accurate text and table extraction
from app.utils.PdfOcrService import PDFOCRService, OCRError
import openpyxl  # XLSX
from langchain_text_splitters import RecursiveCharacterTextSplitter

from app.core.config import settings
from app.utils.piiMasker import PIIMasker

logger = logging.getLogger(__name__)

# Supported MIME types — must match vaultModel allowlist
MIME_PDF  = "application/pdf"
MIME_CSV  = "text/csv"
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SUPPORTED_MIME_TYPES = {MIME_PDF, MIME_CSV, MIME_XLSX}

# Minimum characters per page from pymupdf before we consider it a scanned page
# and fall back to OCR for that specific page only.
OCR_FALLBACK_MIN_CHARS_PER_PAGE = 50


@dataclass
class TextChunk:
    """A single chunk ready for embedding."""
    text: str
    chunk_index: int
    page_number: Optional[int]  # page for PDF; sheet index for spreadsheets; None for CSV
    source: str                 # original filename


class DocumentService:
    """
    Stateless multi-format document processor.
    Dispatches to the correct extractor based on mimeType,
    then chunks all formats the same way via langchain splitter.

    Dependencies (all in requirements.txt): pymupdf, openpyxl
    """

    _splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.CHUNK_SIZE,
        chunk_overlap=settings.CHUNK_OVERLAP,
        length_function=len,
        separators=["\n\n", "\n", ". ", " ", ""],
    )

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    @classmethod
    def process_vault_document(
        cls,
        base64_data: str,
        original_name: str,
        mime_type: str = MIME_PDF,
        password: str = "",
    ) -> List[TextChunk]:
        """
        Main entry point — replaces PDFService.process_vault_document().
        Accepts PDF, CSV, XLSX.

        Args:
            base64_data:   Vault.data field (base64-encoded file bytes)
            original_name: Vault.originalName (used as source label)
            mime_type:     Vault.mimeType — controls which extractor runs

        Returns:
            List of TextChunk objects ready for embedding
        """
        if mime_type not in SUPPORTED_MIME_TYPES:
            raise ValueError(
                f"Unsupported mimeType '{mime_type}'. "
                f"Supported: {SUPPORTED_MIME_TYPES}"
            )

        try:
            raw_bytes = cls._decode_base64(base64_data)

            if mime_type == MIME_PDF:
                pages = cls._extract_pdf(raw_bytes, original_name, password=password)
                tabular_pages = [p for p in pages if p.get("is_tabular")]
                normal_pages = [p for p in pages if not p.get("is_tabular")]
                chunks = []
                if normal_pages:
                    chunks.extend(cls._chunk_pages(normal_pages, original_name))
                if tabular_pages:
                    chunks.extend(cls._tabular_to_chunks(tabular_pages, original_name))
            elif mime_type == MIME_CSV:
                pages = cls._extract_csv(raw_bytes, original_name)
                chunks = cls._tabular_to_chunks(pages, original_name)
            elif mime_type == MIME_XLSX:
                pages = cls._extract_xlsx(raw_bytes, original_name)
                chunks = cls._tabular_to_chunks(pages, original_name)
            logger.info(
                "Processed '%s' (%s) → %d sections → %d chunks",
                original_name, mime_type, len(pages), len(chunks),
            )
            return chunks

        except Exception as e:
            logger.error("Failed to process '%s' (%s): %s", original_name, mime_type, e)
            raise

    # ------------------------------------------------------------------
    # Base64 decode
    # ------------------------------------------------------------------

    @staticmethod
    def _decode_base64(base64_data: str) -> bytes:
        """Strip data-URI prefix if present, then decode."""
        if "," in base64_data:
            base64_data = base64_data.split(",", 1)[1]
        try:
            return base64.b64decode(base64_data)
        except Exception as e:
            raise ValueError(f"Invalid base64 data: {e}") from e

    # ------------------------------------------------------------------
    # Extractors — each returns List[{page_number, text}]
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_pdf(pdf_bytes: bytes, source: str, password: str = "") -> List[dict]:
        """
        Extract text and structured tables per page from a PDF using pdfplumber.
        """
        pages = []
        ocr_service: Optional[PDFOCRService] = None

        try:
            doc = pdfplumber.open(io.BytesIO(pdf_bytes), password=password)
        except Exception as e:
            if "password" in str(e).lower() or "encrypted" in str(e).lower():
                if password:
                    raise ValueError("PDF_WRONG_PASSWORD")
                else:
                    raise ValueError("PDF_PASSWORD_PROTECTED")
            raise RuntimeError(f"pdfplumber failed to open '{source}': {e}") from e

        try:
            for page_num, page in enumerate(doc.pages):
                tables = page.extract_tables()
                text = page.extract_text() or ""
                
                table_texts = []
                for table in tables:
                    if not table or not table[0]: continue
                    clean_table = [[str(cell).strip().replace('\n', ' ') if cell else "" for cell in row] for row in table]
                    
                    header = " | ".join(clean_table[0])
                    separator = " | ".join(["---"] * len(clean_table[0]))
                    rows = [" | ".join(row) for row in clean_table[1:]]
                    
                    md_table = f"| {header} |\n| {separator} |\n"
                    for row in rows:
                        md_table += f"| {row} |\n"
                        
                    table_texts.append(md_table)
                    
                combined_text = text + "\n\n" + "\n\n".join(table_texts)
                combined_text = combined_text.strip()

                if len(combined_text) >= OCR_FALLBACK_MIN_CHARS_PER_PAGE:
                    is_tabular = len(tables) > 0
                    pages.append({"page_number": page_num + 1, "text": combined_text, "is_tabular": is_tabular})
                    continue

                logger.info("Page %d of '%s' has little text — attempting OCR fallback", page_num + 1, source)
                
                if ocr_service is None:
                    try:
                        ocr_service = PDFOCRService()
                    except Exception as ocr_init_err:
                        logger.warning("OCR service unavailable: %s", ocr_init_err)
                        if combined_text:
                            pages.append({"page_number": page_num + 1, "text": combined_text, "is_tabular": False})
                        continue

                try:
                    fitz_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                    if fitz_doc.needs_pass:
                        fitz_doc.authenticate(password)
                        
                    single_page_doc = fitz.open()
                    single_page_doc.insert_pdf(fitz_doc, from_page=page_num, to_page=page_num)
                    single_page_bytes = single_page_doc.tobytes()
                    single_page_doc.close()
                    fitz_doc.close()

                    ocr_text = ocr_service.extract_text_from_bytes(
                        single_page_bytes,
                        filename=f"{source}_page{page_num + 1}.pdf",
                    ).strip()

                    if ocr_text:
                        pages.append({"page_number": page_num + 1, "text": ocr_text, "is_tabular": False})
                    else:
                        if combined_text:
                            pages.append({"page_number": page_num + 1, "text": combined_text, "is_tabular": False})

                except Exception as ocr_err:
                    logger.warning("OCR failed: %s", ocr_err)
                    if combined_text:
                        pages.append({"page_number": page_num + 1, "text": combined_text, "is_tabular": False})

        finally:
            doc.close()
            if ocr_service is not None:
                ocr_service.close()

        if not pages:
            raise ValueError(f"No extractable text found in '{source}'.")
        return pages

    @staticmethod
    def _extract_csv(csv_bytes: bytes, source: str) -> List[dict]:
        """
        Parse CSV into a single text block formatted as a Markdown table.
        page_number is None (no page concept for CSV).
        """
        try:
            text_content = csv_bytes.decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text_content))
            rows = list(reader)
            if not rows:
                raise ValueError(f"CSV '{source}' is empty")

            lines = []
            headers = [str(h).strip().replace('|', '-') for h in rows[0]]
            if headers:
                lines.append("| " + " | ".join(headers) + " |")
                lines.append("|" + "|".join(["---"] * len(headers)) + "|")
                
            for row in rows[1:]:
                clean_row = [str(cell).strip().replace('|', '-') for cell in row]
                # Ensure row matches header length
                padded_row = clean_row + [""] * (len(headers) - len(clean_row))
                padded_row = padded_row[:len(headers)]
                if any(clean_row):
                    lines.append("| " + " | ".join(padded_row) + " |")

            if len(lines) <= 2:
                raise ValueError(f"No data rows found in CSV '{source}'")

            return [{"page_number": None, "text": "\n".join(lines)}]

        except Exception as e:
            raise RuntimeError(f"CSV parsing failed for '{source}': {e}") from e

    @staticmethod
    def _extract_xlsx(xlsx_bytes: bytes, source: str) -> List[dict]:
        """
        Parse XLSX — one section per sheet, formatted as a Markdown table.
        page_number = sheet index (1-based).
        """
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(xlsx_bytes), read_only=True, data_only=True
            )
            pages = []
            for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
                ws = wb[sheet_name]
                lines = []
                is_first_row = True
                header_len = 0
                
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip().replace('|', '-') if c is not None else "" for c in row]
                    
                    if not any(cells):
                        continue
                        
                    if is_first_row:
                        header_len = len(cells)
                        lines.append(f"| " + " | ".join(cells) + " |")
                        lines.append("|" + "|".join(["---"] * header_len) + "|")
                        is_first_row = False
                    else:
                        padded_cells = cells + [""] * (header_len - len(cells))
                        padded_cells = padded_cells[:header_len]
                        lines.append(f"| " + " | ".join(padded_cells) + " |")
                        
                if lines:
                    pages.append({
                        "page_number": sheet_idx,
                        "text": f"[Sheet: {sheet_name}]\n\n" + "\n".join(lines),
                    })
            wb.close()

            if not pages:
                raise ValueError(f"No data found in XLSX '{source}'")
            return pages

        except Exception as e:
            raise RuntimeError(f"XLSX parsing failed for '{source}': {e}") from e

    # ------------------------------------------------------------------
    # Chunker — shared by all formats
    # ------------------------------------------------------------------

    @classmethod
    def _tabular_to_chunks(cls, pages: List[dict], source: str) -> List[TextChunk]:
        """
        Implements dynamic 2x chunking for tables.
        If a table is larger than 2x the standard chunk size, we split it while
        attaching the header to every chunk.
        """
        chunks: List[TextChunk] = []
        chunk_index = 0
        
        max_table_chunk_size = settings.CHUNK_SIZE * 2

        for page in pages:
            masked_text, findings = PIIMasker.mask_with_report(page["text"])
            
            lines = masked_text.split('\n')
            if not lines:
                continue
                
            header = ""
            start_idx = 0
            for i, line in enumerate(lines):
                if line.strip() and not line.startswith("[Sheet:") and not line.startswith("| ---"):
                    header = line
                    start_idx = i + 1
                    # Skip the separator line in markdown tables
                    if len(lines) > start_idx and lines[start_idx].startswith("| ---"):
                        start_idx += 1
                    break
            
            if len(masked_text) <= max_table_chunk_size:
                chunks.append(TextChunk(
                    text=masked_text,
                    chunk_index=chunk_index,
                    page_number=page.get("page_number"),
                    source=source,
                ))
                chunk_index += 1
            else:
                current_chunk_lines = [header] if header else []
                current_len = len(header)
                
                for line in lines[start_idx:]:
                    if not line.strip() or line.startswith("| ---"):
                        continue
                        
                    if current_len + len(line) > max_table_chunk_size and len(current_chunk_lines) > 1:
                        chunks.append(TextChunk(
                            text="\n".join(current_chunk_lines),
                            chunk_index=chunk_index,
                            page_number=page.get("page_number"),
                            source=source,
                        ))
                        chunk_index += 1
                        current_chunk_lines = [header] if header else []
                        current_len = len(header)
                        
                    current_chunk_lines.append(line)
                    current_len += len(line) + 1
                    
                if len(current_chunk_lines) > (1 if header else 0):
                    chunks.append(TextChunk(
                        text="\n".join(current_chunk_lines),
                        chunk_index=chunk_index,
                        page_number=page.get("page_number"),
                        source=source,
                    ))
                    chunk_index += 1



        return chunks

    @classmethod
    def _chunk_pages(cls, pages: List[dict], source: str) -> List[TextChunk]:
        """
        Chunk each page/section and preserve page_number mapping.
        Uses langchain RecursiveCharacterTextSplitter.
        PII is masked before chunking so no sensitive data is embedded.
        """
        chunks: List[TextChunk] = []
        chunk_index = 0

        for page in pages:
            # Mask PII before splitting — keeps placeholders intact across chunks
            masked_text, findings = PIIMasker.mask_with_report(page["text"])
            if findings:
                logger.info(
                    "🔒 PII masked in '%s' page %s — %s",
                    source,
                    page.get("page_number", "N/A"),
                    ", ".join(f"{f.pii_type}×{f.count}" for f in findings),
                )
            page_chunks = cls._splitter.split_text(masked_text)
            for raw_chunk in page_chunks:
                cleaned = raw_chunk.strip()
                if not cleaned:
                    continue
                chunks.append(TextChunk(
                    text=cleaned,
                    chunk_index=chunk_index,
                    page_number=page["page_number"],
                    source=source,
                ))
                chunk_index += 1

        return chunks